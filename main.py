# Import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *

# Opening the existing excel file
wb = load_workbook('/Users/jvpurushotham/Downloads/Registration_form.xlsx')

# Create the sheet object
sheet = wb.active

def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Course"
    sheet.cell(row=1, column=3).value = "Semester"
    sheet.cell(row=1, column=4).value = "Form Number"
    sheet.cell(row=1, column=5).value = "Contact Number"
    sheet.cell(row=1, column=6).value = "Email id"
    sheet.cell(row=1, column=7).value = "Address"

def focus1(event):
    course_field.focus_set()

def focus2(event):
    sem_field.focus_set()

def focus3(event):
    form_no_field.focus_set()

def focus4(event):
    contact_no_field.focus_set()

def focus5(event):
    email_id_field.focus_set()

def focus6(event):
    address_field.focus_set()

def clear():
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)

def insert():
    if (name_field.get() == "" and 
        course_field.get() == "" and
        sem_field.get() == "" and
        form_no_field.get() == "" and
        contact_no_field.get() == "" and
        email_id_field.get() == "" and
        address_field.get() == ""):
        print("empty input")
    else:
        current_row = sheet.max_row
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()

        # Save the file
        wb.save('/Users/jvpurushotham/Downloads/Registration_form.xlsx')

        name_field.focus_set()
        clear()

if __name__ == "__main__":
    root = Tk()
    root.configure(background='#f0f8ff')  # Light blue background
    root.title("Registration Form")

    # Set the window size and position it in the center of the screen
    window_width, window_height = 600, 500
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    position_top = int((screen_height / 2) - (window_height / 2))
    position_right = int((screen_width / 2) - (window_width / 2))
    root.geometry(f"{window_width}x{window_height}+{position_right}+{position_top}")

    excel()

    # Define styling for labels
    label_style = {'bg': '#f0f8ff', 'fg': '#333333', 'font': ('Arial', 12, 'bold')}

    heading = Label(root, text="Registration Form", bg='#4682b4', fg='white', font=('Arial', 18, 'bold'), pady=10)
    heading.grid(row=0, column=1, columnspan=2, pady=10)

    name = Label(root, text="Name", **label_style)
    course = Label(root, text="Course", **label_style)
    sem = Label(root, text="Semester", **label_style)
    form_no = Label(root, text="Form No.", **label_style)
    contact_no = Label(root, text="Contact No.", **label_style)
    email_id = Label(root, text="Email ID", **label_style)
    address = Label(root, text="Address", **label_style)

    name.grid(row=1, column=0, padx=10, pady=5, sticky=E)
    course.grid(row=2, column=0, padx=10, pady=5, sticky=E)
    sem.grid(row=3, column=0, padx=10, pady=5, sticky=E)
    form_no.grid(row=4, column=0, padx=10, pady=5, sticky=E)
    contact_no.grid(row=5, column=0, padx=10, pady=5, sticky=E)
    email_id.grid(row=6, column=0, padx=10, pady=5, sticky=E)
    address.grid(row=7, column=0, padx=10, pady=5, sticky=E)

    name_field = Entry(root, width=40, font=('Arial', 10))
    course_field = Entry(root, width=40, font=('Arial', 10))
    sem_field = Entry(root, width=40, font=('Arial', 10))
    form_no_field = Entry(root, width=40, font=('Arial', 10))
    contact_no_field = Entry(root, width=40, font=('Arial', 10))
    email_id_field = Entry(root, width=40, font=('Arial', 10))
    address_field = Entry(root, width=40, font=('Arial', 10))

    name_field.bind("<Return>", focus1)
    course_field.bind("<Return>", focus2)
    sem_field.bind("<Return>", focus3)
    form_no_field.bind("<Return>", focus4)
    contact_no_field.bind("<Return>", focus5)
    email_id_field.bind("<Return>", focus6)

    name_field.grid(row=1, column=1, padx=10, pady=5, ipady=5)
    course_field.grid(row=2, column=1, padx=10, pady=5, ipady=5)
    sem_field.grid(row=3, column=1, padx=10, pady=5, ipady=5)
    form_no_field.grid(row=4, column=1, padx=10, pady=5, ipady=5)
    contact_no_field.grid(row=5, column=1, padx=10, pady=5, ipady=5)
    email_id_field.grid(row=6, column=1, padx=10, pady=5, ipady=5)
    address_field.grid(row=7, column=1, padx=10, pady=5, ipady=5)

    submit = Button(root, text="Submit", fg="black", bg="#4682b4", font=('Arial', 12, 'bold'), command=insert)
    submit.grid(row=8, column=1, pady=20, ipadx=10)

    root.mainloop()
